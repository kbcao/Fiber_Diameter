import numpy as np
import tensorflow as tf

np.random.seed(1337)  # for reproducibility
from keras.datasets import mnist
from keras.utils import np_utils
from keras.models import Sequential
from keras.layers import Dense, Activation, Convolution2D, MaxPooling2D, Flatten, Dropout
from keras.optimizers import Adam
from keras.callbacks import Callback
from keras.callbacks import EarlyStopping
from keras.models import load_model
from openpyxl import load_workbook
from numpy import array
from sklearn import preprocessing
import matplotlib.pyplot as plt

excelFile = 'Data/61440_data.xlsx'  # 要打开的excel文件名
sheetName = 'Sheet1'  # 要读取数据的工作簿名称

ratio_of_training_data = 0.8  # 设置训练集占全部数据的比例
cnn_size = 8
time_step = cnn_size * cnn_size
train_apochs = 10
is_training = 1
is_scale = 1
batch_size=1

wb = load_workbook(excelFile)
ws = wb.get_sheet_by_name(sheetName)

# X
list1 = list()
for row in ws.iter_rows():
    try:
        if row[0].value is not None and row[1].value is not None and row[2].value is not None:
            x = float(row[0].value)
            y = float(row[1].value)
            z = float(row[2].value)
    except ValueError:
        continue
    if x != 0 and y != 0 and z != 0:
        list1.append([x, y, z])

# Y
list2 = list()
for row in ws.iter_rows():
    try:
        if row[2].value is not None:
            z = float(row[2].value)
    except ValueError:
        continue
    if z != 0:
        list2.append(z)

vec1 = array(list1, ndmin=4)
vec2 = array(list2, ndmin=2)

__form_datax_unscale = np.reshape(vec1, [-1, 3])
__form_datay_unscale = np.reshape(vec2, [-1, 1])
__all_data_x_unscale, __all_data_y_unscale = [], []
for i in range(len(__form_datax_unscale) - time_step - 1):
    x = __form_datax_unscale[i:i + time_step]
    y = __form_datay_unscale[i + time_step]
    __all_data_x_unscale.append(x.tolist())
    __all_data_y_unscale.append(y.tolist())

__temp_all_x = np.reshape(__all_data_x_unscale, [-1, cnn_size, cnn_size, 3])
__temp_all_y = np.reshape(__all_data_y_unscale, [-1, 1])

__temp_amount_x = int(len(__temp_all_x) * ratio_of_training_data)
__temp_amount_y = int(len(__temp_all_y) * ratio_of_training_data)

test_data_x_unscale = __temp_all_x[__temp_amount_x:]
test_data_y_unscale = __temp_all_y[__temp_amount_y:]

if is_scale:
    list1_scale = preprocessing.scale(list1)
    list2_scale = preprocessing.scale(list2)
    vec1 = array(list1_scale, ndmin=4)
    vec2 = array(list2_scale, ndmin=2)

form_datax = np.reshape(vec1, [-1, 3])
form_datay = np.reshape(vec2, [-1, 1])
all_data_x, all_data_y = [], []
for i in range(len(form_datax) - time_step - 1):
    x = form_datax[i:i + time_step]
    y = form_datay[i + time_step]
    all_data_x.append(x.tolist())
    all_data_y.append(y.tolist())

all_data_x = np.reshape(all_data_x, [-1, cnn_size, cnn_size, 3])
all_data_y = np.reshape(all_data_y, [-1, 1])

amount_x = int(len(all_data_x) * ratio_of_training_data)
amount_y = int(len(all_data_y) * ratio_of_training_data)

training_data_x = all_data_x[0:amount_x]
training_data_y = all_data_y[0:amount_y]

test_data_x = all_data_x[amount_x:]
test_data_y = all_data_y[amount_y:]

model = Sequential()
model.add(Convolution2D(  # 32*32*32
    batch_input_shape=(None, cnn_size, cnn_size, 3),
    filters=32,
    kernel_size=5,
    strides=1,
    padding='same',  # Padding method
))
model.add(Activation('relu'))

model.add(MaxPooling2D(  # 32*16*16
    pool_size=2,
    strides=2,
    padding='same',  # Padding method
))
model.add(Dropout(0.02, noise_shape=None, seed=None))

model.add(Convolution2D(  # 64*16*16
    64,
    5,
    strides=1,
    padding='same',
))
model.add(Activation('relu'))

model.add(MaxPooling2D(  # 64*8*8
    2,
    2,
    'same',
))

model.add(Flatten())
model.add(Dense(cnn_size * cnn_size * 2))
model.add(Activation('relu'))

model.add(Dense(cnn_size * cnn_size))
model.add(Activation('linear'))

model.add(Dense(1))
model.add(Activation('linear'))

adam = Adam(lr=1e-4)

# We add metrics to get more results you want to see
model.compile(optimizer=adam,
              loss='mse',
              metrics=['accuracy'])


class LossHistory(Callback):
    def on_train_begin(self, logs={}):
        self.loss = []

    def on_epoch_end(self, batch, logs={}):
        self.loss.append(logs.get('loss'))


history = LossHistory()

early_stopping = EarlyStopping(monitor='loss', patience=5, verbose=0, mode='auto')

if is_training:
    print('Training ------------')
    model.fit(training_data_x, training_data_y, epochs=train_apochs, batch_size=batch_size, verbose=1,
              callbacks=[history, early_stopping], )
    # json_string = model.to_json()
    # open('/var/my_model_architecture.json', 'w').write(json_string)
    # model.save_weights('/var/my_model_weights.h5')
    model.save('modelsave.h5')

else:
    # model = model_from_json(open('my_model_architecture.json').read())
    # model.load_weights('my_model_weights.h5')
    model = load_model('modelsave.h5')
    print('\nTesting ------------')
    loss, accuracy = model.evaluate(test_data_x, test_data_y, batch_size=batch_size, )
    prediction_data = model.predict(test_data_x, batch_size=batch_size)

    print('\ntest loss: ', loss)
    print('\ntest accuracy: ', accuracy)

    prediction_data = prediction_data * np.std(test_data_y_unscale) + np.mean(test_data_y_unscale)


    def accuracy(confidence, test_data, prediction_data):
        count = 0
        for i in range(len(prediction_data)):
            if abs(prediction_data[i] - test_data[i]) <= confidence:
                count = count + 1
        print('%.2f%%' % (count * 100 / len(prediction_data)))


    stat_original = test_data_y_unscale.flatten()
    stat_prediction = prediction_data.flatten()
    accuracy(0.1, stat_original, stat_prediction)
    accuracy(0.01, stat_original, stat_prediction)
    accuracy(0.001, stat_original, stat_prediction)

    plt.figure()
    plt.plot(list(range(len(test_data_y_unscale) )), test_data_y_unscale.flatten(), color='b',
             label='Original data')
    plt.plot(list(range(len(prediction_data) )), prediction_data.flatten(), color='r',
             label='Prediction data')
    plt.xlabel('Time (second)')
    plt.ylabel('Fiber diameter (mm)')
    plt.show()


# plt.plot(range(1,101), history.loss)
# plt.xlabel('Epochs')
# plt.ylabel('loss')
# plt.show()
