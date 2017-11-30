from openpyxl import load_workbook
import tensorflow as tf
from random import choice, shuffle
import numpy as np
from numpy import array
import matplotlib.pyplot as plt

# GlobalParam:
time_step = 40  # 时间步
rnn_unit = 400  # hidden layer units
batch_size = 60  # 每一批次训练多少个样例
input_size = 2  # 输入层维度
output_size = 1  # 输出层维度
lr = 0.0006  # 学习率
excelFile = 'Data/61440_data.xlsx'  # 要打开的excel文件名
sheetName = 'Sheet1'  # 要读取数据的工作簿名称
columnName = 'C'  # 要读取数据的列名称
ratio_of_training_data = 0.75  # 设置训练集占全部数据的比例
loss_error = 0.0015  # 损失函数最大误差
all_data_x = []  # 所有数据
all_data_y = []
training_data_x = []  # 训练数据
training_data_y = []
test_data_x = []  # 测试数据
test_data_y = []
prediction_data = []  # 预测值
is_drop_out = 0  # 是否带dropout
X = tf.placeholder(tf.float32, [None, time_step, input_size])  # 每批次输入网络的tensor
Y = tf.placeholder(tf.float32, [None, time_step, output_size])  # 每批次tensor对应的标签
# 输入层、输出层权重、偏置
weights = {
    'in': tf.Variable(tf.random_normal([input_size, rnn_unit])),
    'out': tf.Variable(tf.random_normal([rnn_unit, 1]))
}
biases = {
    'in': tf.Variable(tf.constant(0.1, shape=[rnn_unit, ])),
    'out': tf.Variable(tf.constant(0.1, shape=[1, ]))
}


def get_training_and_test_data():
    wb = load_workbook(excelFile)
    ws = wb.get_sheet_by_name(sheetName)

    # XXXX
    list1 = list()
    for row in ws.iter_rows():
        try:
            if row[0].value is not None and row[1].value is not None:
                x = float(row[0].value)
                y = float(row[1].value)
        except ValueError:
            continue
        if x != 0 and y != 0:
            list1.append([x, y])

    # YYYY
    list2 = list()
    for row in ws.iter_rows():
        try:
            if row[2].value is not None:
                z = float(row[2].value)
        except ValueError:
            continue
        if z != 0:
            list2.append(z)

    vec1 = array(list1, ndmin=3)
    vec2 = array(list2, ndmin=2)

    global all_data_x, all_data_y, training_data_x, training_data_y, test_data_x, test_data_y

    all_data_x = vec1.reshape([-1,2])
    all_data_y = vec2.transpose()

    amount_x = int(len(all_data_x) * ratio_of_training_data)
    amount_y = int(len(all_data_y) * ratio_of_training_data)

    training_data_x = all_data_x[0:amount_x]
    training_data_y = all_data_y[0:amount_y]
    # print(training_data_x)
    # print(training_data_y)
    test_data_x = all_data_x[amount_x:]
    test_data_y = all_data_y[amount_y:]


def lstm(batch):  # 参数：输入网络批次数目
    w_in = weights['in']
    b_in = biases['in']
    input_x = tf.reshape(X, [-1, input_size])  # 需要将tensor转成2维进行计算，计算后的结果作为隐藏层的输入
    input_rnn = tf.matmul(input_x, w_in) + b_in
    input_rnn = tf.reshape(input_rnn, [-1, time_step, rnn_unit])  # 将tensor转成3维，作为lstm cell的输入

    if is_drop_out == 0:
        cell = tf.contrib.rnn.BasicLSTMCell(rnn_unit)
        init_state = cell.zero_state(batch, dtype=tf.float32)
        output_rnn, final_states = tf.nn.dynamic_rnn(cell, input_rnn, initial_state=init_state,
                                                     dtype=tf.float32)  # output_rnn是记录lstm每个输出节点的结果，final_states是最后一个cell的结果
    else:
        multicell = tf.contrib.rnn.MultiRNNCell([dropout_cell(), create_cell()])
        init_state = multicell.zero_state(batch, dtype=tf.float32)
        output_rnn, final_states = tf.nn.dynamic_rnn(multicell, input_rnn, initial_state=init_state, dtype=tf.float32)

    output = tf.reshape(output_rnn, [-1, rnn_unit])  # 作为输出层的输入
    w_out = weights['out']
    b_out = biases['out']
    pred = tf.matmul(output, w_out) + b_out
    return pred, final_states


def create_cell():
    cell = tf.contrib.rnn.BasicLSTMCell(rnn_unit)
    return cell


def dropout_cell():
    cell = tf.contrib.rnn.BasicLSTMCell(rnn_unit)
    lstm_cell = tf.contrib.rnn.DropoutWrapper(cell, output_keep_prob=0.675)
    return lstm_cell


def train(normalized):
    if normalized == 1:
        normalized_training_data_x = (training_data_x - np.mean(training_data_x)) / np.std(training_data_x)
        normalized_training_data_y = (training_data_y - np.mean(training_data_y)) / np.std(training_data_y)
    else:
        normalized_training_data_x = training_data_x
        normalized_training_data_y = training_data_y

    trainingData_X, trainingData_Y = [], []
    for i in range(len(normalized_training_data_x) - time_step - 1):  # 训练集数据
        x = normalized_training_data_x[i:i + time_step]  # raw data
        y = normalized_training_data_y[i + 1:i + time_step + 1]  # the next second data of raw data
        trainingData_X.append(x.tolist())
        trainingData_Y.append(y.tolist())

    pred, _ = lstm(batch_size)
    # 损失函数
    loss = tf.reduce_mean(tf.square(tf.reshape(pred, [-1]) - tf.reshape(Y, [-1])))
    train_op = tf.train.AdamOptimizer(lr).minimize(loss)
    saver = tf.train.Saver(tf.global_variables())
    with tf.Session() as sess:
        sess.run(tf.global_variables_initializer())
        i = 0
        while i <= 100:  # 最多把训练集遍历100遍
            step = 0
            start = 0
            end = start + batch_size
            i = i + 1
            loss_ = float('inf')  # 无穷大l
            while end < len(trainingData_X) and loss_ > loss_error:  # 重复训练直到遍历所有数据或小余误差允许值Epsilon
                _, loss_ = sess.run([train_op, loss],
                                    feed_dict={X: trainingData_X[start:end], Y: trainingData_Y[start:end]})
                start += batch_size
                end = start + batch_size
                step += 1
                if loss_ <= loss_error:
                    print(i, step, loss_)
                    print("model save：", saver.save(sess, 'result\lstm'))
                    return
                # 每100步显示一次参数，防止误以为死机
                if step % 10 == 0:
                    print('i=', i, 'step=', step, 'loss=', loss_)
        print('no session is recorded. please increase the loss_error value and try again')


def predict(normalized):
    global prediction_data
    if normalized == 1:
        normalized_test_data_x = (test_data_x - np.mean(test_data_x)) / np.std(test_data_x)
        normalized_test_data_y = (test_data_y - np.mean(test_data_y)) / np.std(test_data_y)
    else:
        normalized_test_data_x = test_data_x
        normalized_test_data_y = test_data_y

    testData_X, testData_Y = [], []
    for i in range(len(normalized_test_data_x) - time_step - 1):  # 训练集数据
        x = normalized_test_data_x[i:i + time_step]  # raw data
        y = normalized_test_data_y[i + 1:i + time_step + 1]  # the next second data of raw data
        testData_X.append(x.tolist())
        testData_Y.append(y.tolist())

    pred, _ = lstm(1)  # input[1,time_step,input_size]
    saver = tf.train.Saver(tf.global_variables())
    sess = tf.Session()
    # parameter fetch
    sess.run(tf.global_variables_initializer())
    saver.restore(sess, 'result\lstm')
    for i in range(len(testData_X)):
        next_seq = sess.run(pred, feed_dict={X: [testData_X[i]], Y: [testData_Y[i]]})
        prediction_data.append(next_seq[-1])  # 添加最后一个值到预测队列??????
    if normalized == 1:
        for i in range(len(prediction_data)):
            prediction_data[i] = prediction_data[i] * np.std(test_data_y) + np.mean(test_data_y)


def show_plot():
    accuracy(0.1, test_data_y, prediction_data)
    accuracy(0.01, test_data_y, prediction_data)
    accuracy(0.001, test_data_y, prediction_data)
    plt.figure()
    plt.plot(list(range(len(test_data_y[time_step + 1:]))), test_data_y[time_step + 1:], color='b',
             label='Original data')
    plt.plot(list(range(len(prediction_data))), prediction_data, color='r', label='Prediction data')
    plt.legend(loc=1, ncol=1)
    plt.xlabel('Time (second)')
    plt.ylabel('Fiber diameter (mm)')
    # plt.savefig("LSTM.PNG")
    plt.show()


def accuracy(confidence, test_data, prediction_data):
    count = 0
    for i in range(len(prediction_data)):
        if abs(prediction_data[i] - test_data[i + time_step + 1]) <= confidence:
            count = count + 1
    print('%.2f%%' % (count * 100 / len(prediction_data)))


def main():
    get_training_and_test_data()
    is_training = 1  # 是进行训练还是进行预测
    normalized = 1  # 是否进行数据标准化 1 为标准化 0 不要标准化
    if is_training == 1:
        train(normalized)
    else:
        predict(normalized)
        show_plot()


main()

#1. CNN的输出应该是什么（目前是cnn_size*cnn_size）
#2. 为什么预测结果和真实结果总是差一个常数，而loss很小
